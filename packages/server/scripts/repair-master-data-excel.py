#!/usr/bin/env python3
"""
Repair Master Data.xlsx: Excel Tables (auto-expand) + Name Manager — no corruption.

Root cause of Excel repair dialog: defined names MUST NOT share the same name as
a Table displayName (e.g. name "Adhesive" + table "Adhesive" breaks table2.xml).

Tables use tbl* prefix; Name Manager keeps friendly names (Adhesive, Unit, etc.).

Run:
  python packages/server/scripts/repair-master-data-excel.py
  python packages/server/scripts/repair-master-data-excel.py "path/to/Master Data.xlsx"
"""
from __future__ import annotations

import re
import shutil
import sys
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

REPO_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_PATH = REPO_ROOT / "Master Data.xlsx"

TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)

# sheet_name -> (table_display_name, header_fixes)
SHEETS: dict[str, tuple[str, dict[str, str]]] = {
    "Substrate": ("tblSubstrates", {"Market Price ": "Market Price"}),
    "Ink & Coating": ("tblInkCoating", {}),
    "Adhesive": ("tblAdhesive", {}),
    "Packaging": ("tblPackaging", {}),
    "Unit": ("tblUnits", {}),
    "PT": ("tblPtypes", {}),
    "RM Type": ("tblRMTypes", {}),
    "Printing Web": ("tblPrintingWeb", {}),
}

PT_CODE_BY_LABEL = {
    "Roll": "roll",
    "Sleeve": "sleeve",
    "Bag": "pouch",
    "Pouch": "pouch",
    "Bag or Pouch": "pouch",
}

PRINTING_WEB_DEFAULTS = [
    ["Printing Web", "Code", "Ink System", "Solid %"],
    ["Wide Web", "wide_web", "Ink SB", 30],
    ["Narrow Web", "narrow_web", "Ink UV", 100],
]

# Friendly Name Manager name -> structured reference (uses tbl* tables)
DEFINED_NAMES: dict[str, str] = {
    "SubstrateFamily": "tblSubstrates[Substrate Family]",
    "BOPP_Transparent": "tblSubstrates[Substrate Grade]",
    "InkCoating": "tblInkCoating[Grade]",
    "Adhesive": "tblAdhesive[Grade]",
    "Packaging": "tblPackaging[Grade]",
    "Unit": "tblUnits[Units]",
    "Ptypes": "tblPtypes[Product Type]",
    "PrintingWeb": "tblPrintingWeb[Printing Web]",
    "Type": "tblRMTypes[RM Type]",
}


def last_data_row(ws, min_col: int = 1) -> int:
    for row in range(ws.max_row, 0, -1):
        val = ws.cell(row, min_col).value
        if val not in (None, ""):
            return row
    return 1


def fix_headers(ws, renames: dict[str, str]) -> None:
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val is None:
            continue
        text = str(val).strip()
        for old, new in renames.items():
            if str(val) == old or text == old.strip():
                ws.cell(1, col).value = new
                break


def clear_tables(ws) -> None:
    for name in list(ws.tables.keys()):
        del ws.tables[name]


def add_table(ws, display_name: str) -> str:
    last_row = last_data_row(ws)
    last_col = ws.max_column
    # Trim trailing empty columns
    while last_col > 1:
        empty = all(
            ws.cell(r, last_col).value in (None, "")
            for r in range(1, last_row + 1)
        )
        if not empty:
            break
        last_col -= 1

    col_letter = openpyxl.utils.get_column_letter(last_col)
    ref = f"A1:{col_letter}{last_row}"

    clear_tables(ws)
    table = Table(displayName=display_name, ref=ref)
    table.tableStyleInfo = TABLE_STYLE
    ws.add_table(table)
    return ref


def validate_saved(path: Path, table_names: list[str]) -> list[str]:
    errors: list[str] = []
    with zipfile.ZipFile(path) as z:
        defined = set(re.findall(r'<definedName name="([^"]+)"', z.read("xl/workbook.xml").decode()))
        tbl_xml = b"".join(z.read(n) for n in z.namelist() if n.startswith("xl/tables/"))
        tbl_display = set(re.findall(r'displayName="([^"]+)"', tbl_xml.decode("utf-8")))

        collisions = defined & tbl_display
        if collisions:
            errors.append(f"Name/table collision: {sorted(collisions)}")

        for tn in table_names:
            if tn not in tbl_display:
                errors.append(f"Missing table: {tn}")

        for i in range(1, 20):
            sp = f"xl/worksheets/sheet{i}.xml"
            if sp not in z.namelist():
                continue
            if "<tableParts" not in z.read(sp).decode():
                errors.append(f"{sp} has no tableParts link")

    return errors


def ensure_reference_sheets(wb) -> None:
    """Add Printing Web sheet and PT Code column when missing."""
    if "Printing Web" not in wb.sheetnames:
        ws = wb.create_sheet("Printing Web")
        for row in PRINTING_WEB_DEFAULTS:
            ws.append(row)
        print("  Created sheet: Printing Web (defaults)")

    if "PT" in wb.sheetnames:
        ws = wb["PT"]
        headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        if not any(h.lower() == "code" for h in headers):
            ws.cell(1, ws.max_column + 1).value = "Code"
            label_col = 1
            code_col = ws.max_column
            for row in range(2, ws.max_row + 1):
                label = str(ws.cell(row, label_col).value or "").strip()
                if not label:
                    continue
                existing = str(ws.cell(row, code_col).value or "").strip()
                if not existing:
                    ws.cell(row, code_col).value = PT_CODE_BY_LABEL.get(label, slugify(label))
            print("  PT: added Code column")


def slugify(text: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")
    return s or "value"


def repair_workbook(path: Path) -> None:
    if not path.exists():
        print(f"ERROR: File not found: {path}", file=sys.stderr)
        sys.exit(1)

    backup = path.with_suffix(f".backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx")
    shutil.copy2(path, backup)
    print(f"Backup: {backup}")

    wb = openpyxl.load_workbook(path)

    ensure_reference_sheets(wb)

    for name in list(wb.defined_names.keys()):
        del wb.defined_names[name]

    created_tables: list[str] = []
    print("\nTables:")
    for sheet_name, (table_name, header_fixes) in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  SKIP missing sheet: {sheet_name}")
            continue
        ws = wb[sheet_name]
        fix_headers(ws, header_fixes)
        ref = add_table(ws, table_name)
        created_tables.append(table_name)
        print(f"  {sheet_name} -> {table_name} ({ref})")

    print("\nName Manager:")
    for name, attr_text in DEFINED_NAMES.items():
        wb.defined_names.add(DefinedName(name=name, attr_text=attr_text))
        print(f"  {name} = {attr_text}")

    wb.save(path)

    errors = validate_saved(path, created_tables)
    if errors:
        print("\nVALIDATION FAILED:", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        print(f"Restoring backup to {path}", file=sys.stderr)
        shutil.copy2(backup, path)
        sys.exit(1)

    print(f"\nSaved: {path}")
    print("Validation OK — open in Excel; no repair dialog expected.")
    print("Add rows below any table; Excel expands tbl* range automatically.")


if __name__ == "__main__":
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_PATH
    repair_workbook(target)
